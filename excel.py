from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.dialects.postgresql import insert

from database import SessionLocal
from models import Player

def export_excel(joueurs, fichier):
    wb = Workbook()
    ws = wb.active
    ws.title = "Licenciés"

    entete = [
        "N° licence",
        "Nom",
        "Prénom",
        "Type certificat médical",
        "Type",
        "Catégorie",
        "Points",
        "Validation",
        "Mutation",
    ]

    type_licence = [
        "A = Dirigeant",
        "T = Compétition",
        "P = Loisir",
    ]

    certificat = [
        "P = Parcours Prévention Santé",
        "N = Sans Pratique Sportive",
        "C = Standard Certificat Médical"
        "U = Attestation Autoquestionnaire Pour Mineure",
    ]

    # En-têtes principaux A à I
    for col, titre in enumerate(entete, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = titre
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

    # Explications en colonne K : Type de licence
    ws["K1"] = "Type de licence"
    ws["K1"].font = Font(bold=True)
    ws["K1"].fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    for i, explication in enumerate(type_licence, start=2):
        ws.cell(row=i, column=11).value = explication

    # Explications en colonne L : Certificat médical
    ws["L1"] = "Type certificat médical"
    ws["L1"].font = Font(bold=True)
    ws["L1"].fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    for i, explication in enumerate(certificat, start=2):
        ws.cell(row=i, column=12).value = explication

    # Données des joueurs
    ligne = 2

    for j in joueurs:
        ws.cell(ligne, 1).value = j.licence
        ws.cell(ligne, 2).value = j.nom
        ws.cell(ligne, 3).value = j.prenom
        ws.cell(ligne, 4).value = j.certif
        ws.cell(ligne, 5).value = j.type
        ws.cell(ligne, 6).value = j.categorie
        ws.cell(ligne, 7).value = j.points
        ws.cell(ligne, 8).value = j.validation
        ws.cell(ligne, 9).value = j.mutation

        ligne += 1

    # Ajustement automatique de la largeur des colonnes
    for colonne in ws.columns:
        longueur = max(
            len(str(cell.value)) if cell.value else 0
            for cell in colonne
        )

        ws.column_dimensions[
            colonne[0].column_letter
        ].width = longueur + 3

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(fichier)

def export_neon(joueurs):
    """
    Exporte les joueurs dans la base Neon.

    Si la licence existe déjà :
        -> mise à jour du joueur

    Si la licence n'existe pas :
        -> création du joueur
    """

    donnees = []

    for j in joueurs:
        donnees.append({
            "license": j.licence,
            "name": f"{j.nom} {j.prenom}",
            "ranking": int(j.points or 0),
            "team": j.club,
            "type_certif": j.certif,
            "type_licence": j.type,
            "validation": j.validation,
        })

    if not donnees:
        return 0

    session = SessionLocal()

    try:
        requete = insert(Player).values(donnees)

        requete = requete.on_conflict_do_update(
            index_elements=["license"],
            set_={
                "name": requete.excluded.name,
                "ranking": requete.excluded.ranking,
                "team": requete.excluded.team,
                "type_certif": requete.excluded.type_certif,
                "type_licence": requete.excluded.type_licence,
                "validation": requete.excluded.validation,
            },
        )

        session.execute(requete)
        session.commit()

        return len(donnees)

    except Exception:
        session.rollback()
        raise

    finally:
        session.close()
        
    